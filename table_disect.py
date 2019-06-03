import os
from openpyxl import Workbook, load_workbook

fileName = os.path.join('test.xlsx') # compared file
workbook = load_workbook(fileName)
firstWorkSheet = workbook[workbook.sheetnames[0]]
secondWorkSheet = workbook[workbook.sheetnames[1]]
PRIMARY_KEY = firstWorkSheet['A1'].value # if the primary key is located within the first cell

un_equal_list = []
key_list = []
first_sheet_new_data = []

# Str -> Str
# Creates a new directory with a given name
# and returns its path
# if the directory already exists raise an error
def createNewDir(dirName):
    currentPath =  r'C:\Users\shaiis\Desktop\table_operator' + "\\" + dirName
    if not os.path.exists(currentPath):
        os.makedirs(currentPath)
        return currentPath
    else:
        raise FileExistsError("Give the directory a new name, current directory already exists")

# excelSheet -> Str
# takes a excel work sheet object and returns the sheet name
def getSheetName(excelSheet):
    return excelSheet.title

# list, WorkSheetObject -> null
# recieves a list of matched data and a work sheet object
# produces a table in the given worksheet parameter
def createTable(matchedData, WorkSheetObject, excelSheet):
    WorkSheetObject.append(list(setDictColumns(excelSheet).keys()))

    for row in matchedData:
        WorkSheetObject.append(list(row.values()))

def createWorkSheet(WorkBookObject, sheet, nameStringAddition=''):
    newWorkSheet = WorkBookObject.active
    newWorkSheet.title = getSheetName(sheet) + nameStringAddition

    return newWorkSheet


# Str Str list list ->
def createWorkBookEqual(excelS, matchedList, path):

    WorkBook = Workbook()

    if len(matchedList) != 0:
       equalWorkSheet = createWorkSheet(WorkBook, excelS, 'equal_data_rows')
       createTable(matchedList, equalWorkSheet, excelS)  # create the table inside the work book
       WorkBook.save(path + "\\" + 'equal_data_rows.xlsx')  # save the new work book


# Str  list, Str -> none
def createWorkBookNew(excelS, newList, path):

    WorkBook = Workbook()

    if len(newList) != 0:
       newDataWorkSheet = createWorkSheet(WorkBook, excelS, 'new_data_rows')
       createTable(newList, newDataWorkSheet, excelS)
       WorkBook.save(path + "\\" + 'new_data_rows.xlsx')  # save the new work book

# dict list  Str-> dict
# takes a dict and changes it in place based on the k_list values
# they represent the column for each unequal data
# the function should return the same dict changed, with indicators
# where the dict has equal values and be unchanged where the dict has
# unequal values
# 1. loop through k_list
# 2.      loop through unequal_dict keys
# 3.           if the value in k_list equals the key in unequal_dict
# 4.              break
# 5.           if the value in k_list not equals the key in unequal_dict
# 6.              change unequal_dict[key] in place to "equal_data"
# 7. return the changed unequal_dict
def key_to_unequal_matcher(unequal_dict, k_list, primary_key):

    dict_that_updates = {}
    previous_unequal_value = ""
    primaryKeyValue = unequal_dict[primary_key]

    for key in k_list:
        for key_dict in unequal_dict.keys():
            if key != key_dict and not key_dict in k_list:
               dict_that_updates[key_dict] = "Equal_Data"
            if key == key_dict:
               dict_that_updates[key_dict] = unequal_dict[key_dict]
               previous_unequal_value = key_dict

    dict_that_updates.update({primary_key: unequal_dict[primary_key]})
    unequal_dict.update(dict_that_updates)
    unequal_dict.update({primary_key: primaryKeyValue})

    return unequal_dict


#list list -> list
def edit_out_equal_data(un_EqualList_element, k_list_element, primary_key):

    edited_un_EqualList = []
    counter = 0

    for un_equal_dict in un_EqualList_element:
        edited_un_EqualList.append(key_to_unequal_matcher(un_equal_dict, k_list_element[counter], primary_key))
        counter += 1
        continue

    #print(edited_un_EqualList)
    return edited_un_EqualList


# Str  list list Str Str-> none
def createWorkBookUnEqual(unEqualList, k_list, excelS, path, primary_key):

    WorkBook = Workbook()

    if len(unEqualList) != 0:
       unequalDataWorkSheet = createWorkSheet(WorkBook, firstWorkSheet, '_unequal_data_rows')
       createTable(edit_out_equal_data(unEqualList, k_list, primary_key), unequalDataWorkSheet, excelS)
       WorkBook.save(path + "\\" + 'un_equal_rows.xlsx')  # save the new work book

# Dict, Dict, str -> Dict
# returns a dict full of data if,
# both rows are equal

# check if both primary keyes in row are of equal value
# if they are equal
# continue to run the remainder of the bodies code
# else (if the primary keys are not equal that means the rows are not
# meant to be checked for equality)
# return from the function

# loop through the first dicts keys
# if if the value in the first dict equals the value in the second dict
# initialize the value of equalRows to True
# else (current values) are not equal
# initialize the value of equalRows to False

# if equalRows is true at the end of the equality check
# return the equal data dict
# else (data rows are not equal)
# return false
def checkRowEquality(firstDataDict, secondDataDict, primaryKey):
    equalRows = False
    unEqualCount = 0
    temp_list = []

    if primaryKey not in firstDataDict.keys() and primaryKey not in secondDataDict.keys():
        raise KeyError("Primary key does not match the one within the excel sheet!")
        return

    if firstDataDict[primaryKey] == secondDataDict[primaryKey]:
       primaryKeyEqual = True
       for data in firstDataDict:
           if firstDataDict[data] == secondDataDict[data]:
              equalRows = True
           if firstDataDict[data] != secondDataDict[data]:
              temp_list.append(data)
              unEqualCount += 1
    else:
        return equalRows

    if len(temp_list) != 0:
       key_list.append(temp_list)

    if equalRows == True and unEqualCount == 0:
       return equalRows
    # we can handle the whole data dict since it has unequal rows inside of it
    if unEqualCount != 0:
       equalRows = False
       un_equal_list.append(firstDataDict)
       return equalRows

#
def is_new_value(current_checked_value, checked_sheet_data, pk):

    for data_row in checked_sheet_data:
        print(data_row.values())
        if current_checked_value == data_row[pk]:
           return False
    else:
        return True

# dict dict list list -> Boolean
# if dict row from first sheet is not in second sheet
#    append that row to the new data list
# if dict row from second sheet is not in first sheet
#    append that row to the new data list
# if dict row from first sheet is in second sheet and dict row from second sheet is in first sheet
#    return true (ie, the record exists in both sheets)
# else (if there is one row that doesnt exist in the first sheet and a second row that doesnt exist in the first sheet)
#     append both rows to the new data rows
def check_for_new_data_rows(sheet_one_data, sheet_two_data, pk):


    if len(sheet_one_data) == len(sheet_two_data):


        for data_row in sheet_two_data:
            if is_new_value(data_row[pk], sheet_one_data, pk):
               first_sheet_new_data.append(data_row)
            else:
                continue

        for data_row in sheet_one_data:
            if is_new_value(data_row[pk], sheet_two_data, pk):
               first_sheet_new_data.append(data_row)
            else:
                continue

        #first_sheet_new_data.extend(sheet_one_data)
        return

    if len(sheet_one_data) > len(sheet_two_data):

       for data_row in sheet_two_data:
           if is_new_value(data_row[pk], sheet_one_data, pk):
              first_sheet_new_data.append(data_row)
           else:
              continue

       for data_row in sheet_one_data:
           if is_new_value(data_row[pk], sheet_two_data, pk):
              first_sheet_new_data.append(data_row)
           else:
              continue

       #first_sheet_new_data.extend(sheet_one_data)

       return

    if len(sheet_two_data) > len(sheet_one_data):

        for data_row in sheet_one_data:
            if is_new_value(data_row[pk], sheet_two_data, pk):
                first_sheet_new_data.append(data_row)
            else:
                continue

        for data_row in sheet_two_data:
            if is_new_value(data_row[pk], sheet_one_data, pk):
                first_sheet_new_data.append(data_row)
            else:
                continue

            #first_sheet_new_data.append(data_row)

        return

# excelSheet -> int
# takes an excelsheet and returns the number of rows in that sheet
def get_number_of_rows(excelSheet):

    number_of_rows = 0

    for row in excelSheet.iter_rows(min_row=1, values_only=True):
        number_of_rows += 1

    return number_of_rows

# excelSheet -> int
# takes an excelsheet and returns the number of columns in that sheet
def get_number_of_columns(excelSheet):

    number_of_columns = 0

    for col in excelSheet.iter_cols(min_row=1, values_only=True):
        number_of_columns += 1

    return number_of_columns

# List, List, Str -> list
# will return a list of equal data rows 
# Loop through the first table untill every primary key is proccessed
# if primary key index equals in both worksheets
# check the tuples if they are equal
# specify the table name from which these values are missing
# print these values out in red
# else
# keep looking in the other worksheet for a row with the same primary key
# firstList need to also check if there is no other row that has the same
# key in the other list and handle it, which means its a new data row
# check_both_records_exist() will keep running based on the number
# of rows, needs to stop after proccessing that new data row
def matcher(firstSheetData, secondSheetData, primaryKey, path):

    check_for_new_data_rows(firstSheetData.copy(), secondSheetData.copy(), primaryKey) #todo: designe a new function for proccessing new data

    matchedList = []

    for dataRow in firstSheetData:
        for secondDataRow in secondSheetData:
            if checkRowEquality(dataRow, secondDataRow, primaryKey) == False:
               continue
            if checkRowEquality(dataRow, secondDataRow, primaryKey) == True:
               matchedList.append(dataRow)
               secondSheetData.remove(secondDataRow)
               break

    if first_sheet_new_data != 0:
       createWorkBookNew(firstWorkSheet, first_sheet_new_data, path) # first_sheet_new_data

    return matchedList

# print the dictValue list in a good format
# def printTupleList(dictList):
#     for dicts in dictList:
#         for key, value in dicts.items():
#             print("columnName: {} Value: {} ||".format(key, value), end=' ')
#         print()

    # SheetObject -> dict


# returns a new dict object with columns as keys in it, each time
# the funcion is called
# all the columns will be in english
def setDictColumns(excelSheet):
    dataRowdict = {}
    for col in excelSheet.iter_cols(min_row=1, max_col=get_number_of_columns(excelSheet), max_row=1, values_only=True):
        for cell in col:
            dataRowdict[cell] = ''

    return dataRowdict


# Tuple -> dict
# this function adds values to our dict keys
# the function returns the dict object with key and value pairs filled
def setDictValues(rowOfValues):
    dataDict = setDictColumns(firstWorkSheet)
    rowsIsIter = iter(rowOfValues)

    for key in dataDict.keys():
        dataDict[key] = next(rowsIsIter)

    return dataDict

# sheetObjectList, Integer, Integer -> list of tuples
# number of columns and number of rows will most probably
# be equal on both tables so they can be parameters
def createDataList(excelSheet, maxColumns, maxRows):
    listOfDataDicts = []
    ColumnRow = True

    for row in excelSheet.iter_rows(min_row=1, max_col=get_number_of_columns(excelSheet), max_row=get_number_of_rows(excelSheet), values_only=True):
        if ColumnRow == True:  # dont print the columns
           ColumnRow = False
           continue
        listOfDataDicts.append(setDictValues(row))

    return listOfDataDicts


"""
TEST CASES FOR NEW DATA ROWS
----------------------------
1. first table has more rows and has new data rows while,
the second table has less rows and no new data rows -passed

2.second table has more rows and has a new data row while
the first table has less rows and no new data rows - passed

3. both tables dont have new rows passed - passed

4. both tables have new rows but the same number of rows,
each table has different new rows - passed after fix

5. both tables have new rows but first table has more rows - passed after fix

6. both tables have new data rows but the second table has - passed after fix
   more rows
   
   
TEST CASES FOR EQUAL DATA ROWS
------------------------------
1. both tables have the same amount of rows - passed with one table
having new rows, didnt include them with in the equal sheet 

2. table one has more rows then table two - passed with the first table
having new rows and more rows then the second table

3. table two has more rows then table one - passed with the second table 
having new rows and more rows then the first table which also had mixed 
new rows.


TEST CASES FOR UNEQUAL DATA ROWS
--------------------------------
1. both tables have the same amounth of rows with a few unequal rows injected into
them - passed all though it checks the values as is so some checks might be described as
unequal because of type checking

2. table one has more data then table two, both tables have unequal rows injected into them -
passed

3. table two has more data then table one, both tables have unequal data injected into them -
passed

BUG: WHEN WORKING TOGHETER NEW_DAT AND UNEQUAL_DATA MATCHERS
UNEQUAL_DATA IS BEING WRITTEN AS NEW_DATA

WORK_LOG:
- STORE THE NUMBER OF COLUMNS AUTOMATICALLY - done
- STORE THE NUMBER OF ROWS AUTOMATICALLY - done
- get sheetnames automatically - done
"""
def main():

    listOfDataDictsOne = createDataList(firstWorkSheet, get_number_of_columns(firstWorkSheet),
                                        get_number_of_rows(firstWorkSheet))
    listOfDataDictsTwo = createDataList(secondWorkSheet, get_number_of_columns(secondWorkSheet),
                                        get_number_of_rows(secondWorkSheet))

    path_to_save = createNewDir() # name the directory
    mList = matcher(listOfDataDictsOne, listOfDataDictsTwo, PRIMARY_KEY, path_to_save)

    createWorkBookEqual(firstWorkSheet, mList, path_to_save)
    createWorkBookUnEqual(un_equal_list, key_list, firstWorkSheet, path_to_save, PRIMARY_KEY)

main()
